import openpyxl as pay_xl

pay = pay_xl.load_workbook('c:\\dd\\itx.xlsx')

xl= pay.active

pay_1 = xl[1][1].value
pay_2 = xl[2][1].value
pay_3 = xl[3][1].value
pay_4 = xl[4][1].value
pay_5 = xl[5][1].value

total = pay_1 + pay_2 + pay_3 + pay_4 + pay_5
avg = total / 5

xl.cell(row = 6, column = 3).value = total
xl.cell(row = 6, column = 4).value = avg

pay.save('c:\\dd\\outitx.xlsx')
pay.close()
